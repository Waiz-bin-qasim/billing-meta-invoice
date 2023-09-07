import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Border, Side


def create_excel (filename):
    excel_file = Workbook()
    excel_file.save(filename)



def set_template(sheet,date):

    # fills in blue colour, merges and fills the text
    blue_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    font_style = Font(size=14, bold=True, color='FFFFFF', name='Arial')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    for row in sheet.iter_rows(min_row=1, max_row=3, max_col=6):
        for cell in row:
            cell.fill = blue_fill

    for row_number in range(1,4):
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=6)

    cell = sheet.cell(row=2, column=1, value="Client Wise Billing Report - Digital Connect/WhatsApp API Plan Subscription")
    cell.font = font_style
    cell.alignment = align_center

    font_style = Font(size=12, bold=True, color='FFFFFF', name='Arial')
    cell = sheet.cell(row=3, column=1, value=date)
    cell.font = font_style
    cell.alignment = align_right


    # fills in black colour and text headings
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    font_style = Font(size=10, bold=True, color='FFFFFF', name='Arial')

    for col_num in range(1,7):
        cell = sheet.cell(row=4, column=col_num)
        cell.fill = black_fill

    cell = sheet.cell(row=4, column=2)
    cell.value = "CUSTOMER NAME"
    cell.font = font_style

    cell = sheet.cell(row=4, column=3)
    cell.value = "ITEM"
    cell.font = font_style

    cell = sheet.cell(row=4, column=4)
    cell.value = "BILLLING PERIOD"
    cell.font = font_style

    cell = sheet.cell(row=4, column=5)
    cell.value = "QTY"
    cell.font = font_style
    cell.alignment = align_right

    cell = sheet.cell(row=4, column=6)
    cell.value = "AMOUNT"
    cell.font = font_style
    cell.alignment = align_right


    # sets width of each column
    sheet.column_dimensions['A'].width = 4.22
    sheet.column_dimensions['B'].width = 60.00
    sheet.column_dimensions['C'].width = 50.78
    sheet.column_dimensions['D'].width = 26.89
    sheet.column_dimensions['E'].width = 16.78
    sheet.column_dimensions['F'].width = 38.22



def set_subtotal(sheet,subtotal):

    rowNum = sheet.max_row + 1

    fill_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    align_right = Alignment(horizontal='right', vertical='center')

    for col in range(1,7):
        cell = sheet.cell(row=rowNum, column=col)
        cell.fill = fill_color

    border_style = Side(border_style='thin', color='000000')
    border = Border(top=border_style, bottom=border_style)

    for col in range(1,7):
        cell = sheet.cell(row=rowNum, column=col)
        cell.border = border
    

    font_style = Font(bold=True, name='Arial', size=10)
    cell = sheet.cell(row=rowNum, column=5, value="Subtotal")
    cell.font = font_style
    cell.alignment = align_right

    font_style = Font(bold=True, name='Arial', size=11)
    cell = sheet.cell(row=rowNum, column=6, value=float(subtotal)) 
    cell.font = font_style
    cell.alignment = align_right
    cell.number_format = '$#,##0.00'



def highlight(sheet,colorHEX):

    rowMax = sheet.max_row

    fill_color = PatternFill(start_color=colorHEX, end_color=colorHEX, fill_type='solid')

    for col in range(3,7):
        cell = sheet.cell(row=rowMax, column=col)
        cell.border = None
    
    for col in range(3,7):
        cell = sheet.cell(row=rowMax, column=col)
        cell.fill = fill_color



def planSubscription(sheet,index,company_data):

    font_style = Font(bold=True, name='Arial', size=9)
    rowMax = sheet.max_row + 1
    addIndexRow = 1

    cell = sheet.cell(row=rowMax, column=3, value=company_data[index][9])
    cell.font = font_style
    set_color(sheet)
    highlight(sheet,"B8CCE4")

    # has active users
    if (company_data[index][10] is not None):

        # check PKR OR USD
        if company_data[index][11].upper() == "PKR":
            cell = sheet.cell(row=rowMax+addIndexRow, column=3)
            cell.value = f"{company_data[index][10]}     Monthly Active Users @PKR{company_data[index][12]}/month"
        else:
            cell = sheet.cell(row=rowMax+addIndexRow, column=3)
            cell.value = f"{company_data[index][10]}     Monthly Active Users @${company_data[index][12]}/month"

        cell = sheet.cell(row=rowMax+addIndexRow, column=4)
        cell.value = company_data[index][21] #date

        cell = sheet.cell(row=rowMax+addIndexRow, column=5)
        cell.value = int(company_data[index][13]) #qty

        # check PKR OR USD
        if company_data[index][11].upper() == "PKR":
            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            cell.value = float(company_data[index][14]) #amount
            cell.number_format = 'PKR #,##0.00'
        else:
            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            cell.value = float(company_data[index][14]) #amount
            cell.number_format = '$#,##0.00'

        set_color(sheet)
        highlight(sheet,'D8E4BC')
        addIndexRow = addIndexRow + 1


    # has additional users
    if (company_data[index][15] is not None):

        cell = sheet.cell(row=rowMax+addIndexRow,column=3)
        cell.value = f"Additonal Users Outside Tier @${company_data[index][15]}/per user"

        cell = sheet.cell(row=rowMax+addIndexRow, column=4)
        cell.value = company_data[index][21] #date

        cell = sheet.cell(row=rowMax+addIndexRow,column=5)
        if int(company_data[index][16]) < 0: #qty
            cell.value = 0
        else: 
            cell.value = int(company_data[index][16])

        cell = sheet.cell(row=rowMax+addIndexRow, column=6)
        if int(company_data[index][16]) < 0:
            cell.value = 0
        else:
            cell.value = float(company_data[index][15]) * int(company_data[index][16]) #amount

        cell.number_format = '$#,##0.00'

        set_color(sheet)
        highlight(sheet,'D8E4BC')
        addIndexRow = addIndexRow + 1


    # has agent seats
    if (company_data[index][17] is not None):

        cell = sheet.cell(row=rowMax+addIndexRow, column=3)
        cell.value = f"{company_data[index][17]}     Agent Seats"

        cell = sheet.cell(row=rowMax+addIndexRow, column=4)
        cell.value = company_data[index][21] #date

        cell = sheet.cell(row=rowMax+addIndexRow, column=6)
        cell.value = float(company_data[index][18]) #amount

        cell.number_format = '$#,##0.00'

        set_color(sheet)
        highlight(sheet,'D8E4BC')
        addIndexRow = addIndexRow + 1


    # has platform support 
    if (company_data[index][20] is not None):

        cell = sheet.cell(row=rowMax+addIndexRow, column=3)
        cell.value = f"Platform Support"

        cell = sheet.cell(row=rowMax+addIndexRow, column=4)
        cell.value = company_data[index][21] #date

        # check PKR or USD
        if (company_data[index][19].upper() == 'PKR'):
            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            cell.value = float(company_data[index][20]) #amount
            cell.number_format = 'PKR #,##0.00'
        else:
            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            cell.value = float(company_data[index][20]) #amount
            cell.number_format = '$#,##0.00'

        set_color(sheet)
        highlight(sheet,'D8E4BC')
        addIndexRow = addIndexRow + 1
    

    # sub heading
    cell = sheet.cell(row=rowMax+addIndexRow, column=3, value="WHATSAPP CONVERSATIONS")
    cell.font = font_style
    set_color(sheet)



def set_subscription(sheet,index,company_data):

    font_style = Font(bold=True, name='Arial', size=9)
    rowMax = sheet.max_row
    addIndexRow = 1

    # subscription data
    cell = sheet.cell(row=rowMax, column=3, value="SUBSCRIPTION PLAN")
    cell.font = font_style
    set_color(sheet)

    if (company_data[index][9] is not None):
        planSubscription(sheet,index,company_data)
    else: 

        # has active users
        if (company_data[index][10] is not None):

            # check if PKR OR USD
            if company_data[index][11].upper() == 'PKR': 
                cell = sheet.cell(row=rowMax+addIndexRow, column=3)
                cell.value = f"{company_data[index][10]}     Monthly Active Users @PKR{company_data[index][12]}/month"

            elif company_data[index][11].upper() == 'USD':
                cell = sheet.cell(row=rowMax+addIndexRow, column=3)
                cell.value = f"{company_data[index][10]}     Monthly Active Users @${company_data[index][12]}/month"

            cell = sheet.cell(row=rowMax+addIndexRow, column=4)
            cell.value = company_data[index][21] #date

            cell = sheet.cell(row=rowMax+addIndexRow, column=5)
            cell.value = int(company_data[index][13]) #qty

            # check if PKR OR USD
            if company_data[index][11].upper() == 'PKR': 

                cell = sheet.cell(row=rowMax+addIndexRow, column=6)
                cell.value = float(company_data[index][14]) #amount
                cell.number_format = 'PKR #,##0.00' 

            elif company_data[index][11].upper() == 'USD':
                cell = sheet.cell(row=rowMax+addIndexRow, column=6)
                cell.value = float(company_data[index][14]) #amount
                cell.number_format = '$#,##0.00'

            set_color(sheet)
            addIndexRow = addIndexRow + 1


        # has additional users
        if (company_data[index][15] is not None):

            cell = sheet.cell(row=rowMax+addIndexRow,column=3)
            cell.value = f"Additonal Users Outside Tier @${company_data[index][15]}/per user"

            cell = sheet.cell(row=rowMax+addIndexRow, column=4)
            cell.value = company_data[index][21] #date

            cell = sheet.cell(row=rowMax+addIndexRow,column=5)
            if int(company_data[index][16]) < 0: #qty
                cell.value = 0
            else: 
                cell.value = int(company_data[index][16])

            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            if int(company_data[index][16]) < 0:
                cell.value = 0
            else:
                cell.value = float(company_data[index][15]) * int(company_data[index][16]) #amount

            cell.number_format = '$#,##0.00'
            set_color(sheet)
            addIndexRow = addIndexRow + 1


        # has agent seats
        if (company_data[index][17] is not None):

            cell = sheet.cell(row=rowMax+addIndexRow, column=3)
            cell.value = f"{company_data[index][17]}     Agent Seats"

            cell = sheet.cell(row=rowMax+addIndexRow, column=4)
            cell.value = company_data[index][21] #date

            cell = sheet.cell(row=rowMax+addIndexRow, column=6)
            cell.value = float(company_data[index][18]) #amount

            cell.number_format = '$#,##0.00'
            set_color(sheet)
            addIndexRow = addIndexRow + 1


        # has platform support 
        if (company_data[index][20] is not None):

            cell = sheet.cell(row=rowMax+addIndexRow, column=3)
            cell.value = f"Platform Support"

            cell = sheet.cell(row=rowMax+addIndexRow, column=4)
            cell.value = company_data[index][21] #date

            # check USD or PKR
            if company_data[index][19].upper() == "PKR":
                cell = sheet.cell(row=rowMax+addIndexRow, column=6)
                cell.value = float(company_data[index][20]) #amount
                cell.number_format = 'PKR #,##0.00' 

            elif company_data[index][19].upper() == "USD":
                cell = sheet.cell(row=rowMax+addIndexRow, column=6)
                cell.value = float(company_data[index][20]) #amount
                cell.number_format = '$#,##0.00'

            set_color(sheet)
            addIndexRow = addIndexRow + 1


        # sub heading
        cell = sheet.cell(row=rowMax+addIndexRow, column=3, value="WHATSAPP CONVERSATIONS")
        cell.font = font_style
        set_color(sheet)



def set_heading (sheet,dataNumber,customerName):

    rowMax = sheet.max_row
     
    font_style = Font(bold=True, name='Arial', size=10)
    align_center = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=rowMax+1, column=1, value=dataNumber)
    cell.font = font_style
    cell.alignment = align_center

    cell = sheet.cell(row=rowMax+1, column=2, value=customerName)
    cell.font = font_style

    set_color(sheet)
    
  

def set_footer(sheet):

    rowMax = sheet.max_row
    fill_color = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    font_style = Font(size=10, bold=True, color='000000', name='Arial')
    align_center = Alignment(horizontal='center', vertical='center')

    for row in range(1,7):
        sheet.merge_cells(start_row=rowMax+row, start_column=1, end_row=rowMax+row, end_column=6)
        cell = sheet.cell(row=rowMax+row, column=1)
        cell.fill = fill_color

    cell = sheet.cell(row=rowMax+2, column=1)
    cell.value = "Subtotal for clients does not include any line items  billed in PKR. Hence carefully invoice line items in such cases."
    cell.alignment = align_center
    cell.font = font_style

    cell = sheet.cell(row=rowMax+3, column=1)
    cell.value = "Estimated totals are provided in USD and PKR seperately where applicable."
    cell.alignment = align_center
    cell.font = font_style

    cell = sheet.cell(row=rowMax+4, column=1)
    cell.value = "Estimated totals are only provided for billing items chargeable by Eocean hence WhatsApp conversations fee is not included in Estimated totals."
    cell.alignment = align_center
    cell.font = font_style

    cell = sheet.cell(row=rowMax+5, column=1)
    cell.value = "Verify WhatsApp conversation fee totals from Meta Invoice."
    cell.alignment = align_center
    cell.font = font_style



def set_color(sheet):

    rowMax = sheet.max_row

    fill_color = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for col in range(1,7):
        cell = sheet.cell(row=rowMax, column=col)
        cell.border = None
    
    for col in range(1,7):
        cell = sheet.cell(row=rowMax, column=col)
        cell.fill = fill_color



def insert_data(sheet,company_data):

    # empty array
    if company_data is None or len(company_data) == 0:
        print("EMPTY DATA ARRAY")
        return

    num_rows = len(company_data)
    align_right = Alignment(horizontal='right', vertical='center')
    

    for row in range(num_rows):

        # empty data
        if company_data[row][0] == None:
            continue
    
        
        set_heading(sheet,row+1,company_data[row][0])
        set_subscription(sheet,row,company_data)


        # inserting conversation fee
        rowMax = sheet.max_row
        sheet.cell(row=rowMax+1, column=3, value="Fee Conversation/Month")
        sheet.cell(row=rowMax+1, column=4, value=company_data[row][21]) #date

        cell = sheet.cell(row=rowMax+1, column=5, value=1000) #HARDCODE 1000
        cell.alignment = align_right

        cell = sheet.cell(row=rowMax+1, column=6, value=0.00) #HARDCODE 0.00
        cell.alignment = align_right
        cell.number_format = '$#,##0.00'

        set_color(sheet)


        # inserting Service
        rowMax = sheet.max_row
        sheet.cell(row=rowMax+1, column=3, value="Service Conversation")
        sheet.cell(row=rowMax+1, column=4, value=company_data[row][21]) #date

        cell = sheet.cell(row=rowMax+1, column=5, value=int(company_data[row][1])) #qty
        cell.alignment = align_right

        cell = sheet.cell(row=rowMax+1, column=6, value=float(company_data[row][5])) #amount
        cell.alignment = align_right
        cell.number_format = '$#,##0.00'

        set_color(sheet)

            
        #inserting Marketing
        rowMax = sheet.max_row
        sheet.cell(row=rowMax+1, column=3, value="Marketing Conversation")
        sheet.cell(row=rowMax+1, column=4, value=company_data[row][21]) #date

        cell = sheet.cell(row=rowMax+1, column=5, value=int(company_data[row][2])) #qty
        cell.alignment = align_right

        cell = sheet.cell(row=rowMax+1, column=6, value=float(company_data[row][6])) #amount
        cell.alignment = align_right
        cell.number_format = '$#,##0.00'
      
        set_color(sheet)


        #inserting Utility
        rowMax = sheet.max_row
        sheet.cell(row=rowMax+1, column=3, value="Utility Conversation")
        sheet.cell(row=rowMax+1, column=4, value=company_data[row][21]) #date

        cell = sheet.cell(row=rowMax+1, column=5, value=int(company_data[row][3])) #qty
        cell.alignment = align_right

        cell = sheet.cell(row=rowMax+1, column=6, value=float(company_data[row][7])) #amount
        cell.alignment = align_right
        cell.number_format = '$#,##0.00'

        set_color(sheet)


        #inserting Authentication
        rowMax = sheet.max_row
        sheet.cell(row=rowMax+1, column=3, value="Authentication Conversation")
        sheet.cell(row=rowMax+1, column=4, value=company_data[row][21]) #date

        cell = sheet.cell(row=rowMax+1, column=5, value=int(company_data[row][4])) #qty
        cell.alignment = align_right

        cell = sheet.cell(row=rowMax+1, column=6, value=float(company_data[row][8])) #amount
        cell.alignment = align_right
        cell.number_format = '$#,##0.00'

        set_color(sheet)

        result = calculateSubtotal(company_data,row)

        set_subtotal(sheet,result)
        
    
def calculateSubtotal(company_data,index):

    result = float(company_data[index][22])

    # add monthly price if exists 
    if company_data[index][12] is not None:
        # check if USD
        if company_data[index][11].upper() == "USD":
            result = result + float(company_data[index][14])

    # add additonal user if exists
    if company_data[index][15] is not None:
        # only add if not negative
        if int(company_data[index][16]) > 0:
            result = result + float(company_data[index][15]) * int(company_data[index][16])

    # add agent seats if exists
    if company_data[index][17] is not None:
        result = result + float(company_data[index][18])

    # add platform if exists
    if company_data[index][20] is not None:
        # check if in USD
        if company_data[index][19].upper() == "USD":
            result = result + float(company_data[index][20])
    
    return result



def set_totals(sheet,total_dollars,total_pkr):

    rowMax = sheet.max_row
    align_right = Alignment(horizontal='right', vertical='center')
    font_style = Font(bold=True, name='Arial', size=11)
    border_style = Side(border_style='thin', color='000000')
    border = Border(top=border_style, bottom=border_style)

    for col in range(1,7):
        cell = sheet.cell(row=rowMax+1, column=col)
        cell.border = border
    
    for col in range(1,7):
        cell = sheet.cell(row=rowMax+2, column=col)
        cell.border = border

    # dollars estimate
    sheet.merge_cells(start_row=rowMax+1, start_column=1, end_row=rowMax+1, end_column=5)

    cell = sheet.cell(row=rowMax+1, column=1, value="ESTIMATED TOTAL")
    cell.font = font_style
    cell.alignment = align_right

    cell = sheet.cell(row=rowMax+1, column=6, value=float(total_dollars))
    cell.alignment = align_right
    cell.font = font_style
    cell.number_format = '$#,##0.00'

    # pkr estimate
    sheet.merge_cells(start_row=rowMax+2, start_column=1, end_row=rowMax+2, end_column=5)

    cell = sheet.cell(row=rowMax+2, column=1, value="ESTIMATED TOTAL")
    cell.font = font_style
    cell.alignment = align_right

    cell = sheet.cell(row=rowMax+2, column=6, value=float(total_pkr))
    cell.alignment = align_right
    cell.font = font_style
    cell.number_format = 'PKR #,##0.00'

    

def data_handler(filename,company_data,total_dollars,total_pkr):
   
   try:

        create_excel(filename)

        excel_file = openpyxl.load_workbook(filename)
        sheet = excel_file.active

        set_template(sheet,company_data[0][21]) #date

        insert_data(sheet,company_data)

        set_totals(sheet,total_dollars,total_pkr)

        set_footer(sheet)

        excel_file.save(filename)
        print("DATA SAVED")

   except PermissionError:
        print(f"PermissionError: The file '{filename}' is in use. Please close it and try again.")
        raise Exception(f"PermissionError: The file '{filename}' is in use. Please close it and try again.")
   except Exception as e:
        print(f"Error occurred: {e}")
        raise Exception(e)