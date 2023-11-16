from Config.dbConfig import establish_connection, close_connection
import openpyxl
import os
import PyPDF2
from datetime import datetime, timedelta

def get_previous_month(month, year):
    # Convert input month and year to a datetime object
    current_month = datetime.strptime(f"{year}-{month}-01", "%Y-%m-%d")

def displayTotalClients(month, year):
    try:
        cursor, connection = establish_connection()
        queryClients = (
            "SELECT count(*) FROM MAU WHERE INV_MONTH = %s  AND INV_YEAR = %s"
        )
        cursor.execute(queryClients, (month, year))
        result = cursor.fetchone()
        return result
    except Exception as ex:
        return {"message": "error displaying clients"}


def displayTotalInvoices():
    try:
        cursor, connection = establish_connection()
        queryInvoice = "SELECT count(*) FROM billing.billing_logs"
        cursor.execute(queryInvoice)
        result = cursor.fetchone()
        return result
    except Exception as ex:
        return {"message": "error displaying invoices"}


def displayTotalUSD(month, year):
    # estimated USD to be read from finance report of particular month and year

    try:
        folder = "financeReportFiles"
        fileName = f"{month}{year}.xlsx"
        file_path = os.path.join(folder, fileName)

        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook["Sheet"]

            for row in worksheet.iter_rows(
                min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6
            ):
                if row[0].value and "estimated total" in str(row[0].value).lower():
                    # first estimated total is always dollars
                    return row[5].value

            workbook.close()
        else:
            return 0
    except Exception as ex:
        print(ex)


def displayTotalPKR(month, year):
    # estimated PKR to be read from finance report of particular month and year

    try:
        folder = "financeReportFiles"
        fileName = f"{month}{year}.xlsx"
        file_path = os.path.join(folder, fileName)

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook["Sheet"]

        count = 0
        if os.path.exists(file_path):
            for row in worksheet.iter_rows(
                min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6
            ):
                if row[0].value and "estimated total" in str(row[0].value).lower():
                    count = count + 1
                    # second estimated total is always PKR
                    if count == 2:
                        return row[5].value

        workbook.close()

    except Exception as ex:
        print(ex)


def displayWhatsappAmount(month, year):
    # to be read from meta invoice of particular month and year

    try:
        folder = "metaInvoiceFiles"
        fileName = f"{month}{year}.pdf"
        file_path = os.path.join(folder, fileName)

        if os.path.exists(file_path):
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                page = pdf_reader.pages[0]

                for line in page.extract_text().splitlines():
                    if "Total due by" in line:
                        invoiceTotal = line
                        invoiceTotal = invoiceTotal.split("$")
                        return invoiceTotal[1]

    except Exception as ex:
        print(f"Error while reading PDF: {ex}")


def displayBarChart():
    try:
        mauQuery = "select distinct inv_month,inv_year from mau_logs"
        metaQuery = "select distinct inv_month,inv_year from billing_logs"
        cursor, connection = establish_connection()
        cursor.execute(mauQuery)
        resultMau = cursor.fetchall()
        cursor.execute(metaQuery)
        resultMeta = cursor.fetchall()
        mau_set = {(row[0], row[1]) for row in resultMau}
        meta_set = {(row[0], row[1]) for row in resultMeta}

        totalResult = []
        common_values = mau_set.intersection(meta_set)
        for item in common_values:
            inv_month = item[0]
            inv_year = item[1]
            record = []
            record.append(inv_month)
            record.append(inv_year)
            amount = displayWhatsappAmount(inv_month, inv_year)
            amount = float(amount.replace(",", "").strip())
            record.append(amount)
            record.append(displayTotalUSD(inv_month, inv_year))
            totalResult.append(record)
        return totalResult
    except Exception as ex:
        print(ex)


def get_previous_month(month, year):
    # Convert input month and year to a datetime object
    current_month = datetime.strptime(f"{year}-{month}-01", "%Y-%b-%d")

    # Calculate the first day of the previous month
    previous_month = current_month - timedelta(days=current_month.day)

    # Extract the year and month from the result
    previous_year = previous_month.year
    previous_month = previous_month.strftime('%b')


    return previous_month, previous_year

def data_available(month, year):
    # List of folders where data files may be stored
    data_directories = ["billingMAUFiles", "financeReportFiles", "excel"]

    # Construct the filename based on the month and year
    filename = f"{month}{year}.xlsx"

    # Check if the file exists in any of the data directories
    for data_directory in data_directories:
        file_path = os.path.join(data_directory, filename)
        if os.path.exists(file_path):
            continue
        else:
            return False

    folder = "metaInvoiceFiles"
    filename = f"{month}{year}.pdf"
    file_path = os.path.join(folder, filename)
    if os.path.exists(file_path):
        return True
    else:
        return False
