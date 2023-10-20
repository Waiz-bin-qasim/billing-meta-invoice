from Config.dbConfig import establish_connection, close_connection
import openpyxl
import os
import PyPDF2




def displayTotalClients(month,year):
    try:
        cursor,connection = establish_connection()
        queryClients = "SELECT count(*) FROM MAU WHERE INV_MONTH = %s  AND INV_YEAR = %s"
        cursor.execute(queryClients,(month,year))
        result = cursor.fetchone()
        return result
    except Exception as ex:
        
        return {
            'message' : 'error displaying clients'
        }
        
def displayTotalInvoices():
    try:
        cursor,connection = establish_connection()
        queryInvoice = "SELECT count(*) FROM billing.billing_logs"
        cursor.execute(queryInvoice)
        result = cursor.fetchone()
        return result
    except Exception as ex:
        return {
            'message' : 'error displaying invoices'
        }      
        



def displayTotalUSD(month,year):
    #estimated USD to be read from finance report of particular month and year

    try:

        folder = 'financeReportFiles'
        fileName = f'{month}{year}.xlsx'
        file_path = os.path.join(folder, fileName)


        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook['Sheet']


        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6):
            if row[0].value and 'estimated total' in str(row[0].value).lower():
                # first estimated total is always dollars
                return (row[5].value)


        workbook.close()

    except Exception as ex: 
        print(ex)  


    
def displayTotalPKR(month,year):
    #estimated PKR to be read from finance report of particular month and year

    try:

        folder = 'financeReportFiles'
        fileName = f'{month}{year}.xlsx'
        file_path = os.path.join(folder, fileName)


        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook['Sheet']

        count = 0
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6):
            if row[0].value and 'estimated total' in str(row[0].value).lower():
                count = count + 1
                # second estimated total is always PKR
                if count == 2: 
                    return (row[5].value)


        workbook.close()

    except Exception as ex: 
        print(ex)  



def displayWhatsappAmount(month,year):
    #to be read from meta invoice of particular month and year

    try: 

        
        folder = 'metaInvoiceFiles'
        fileName = f'{month}{year}.pdf'
        file_path = os.path.join(folder, fileName)


        with open(file_path, "rb") as file:
            
            pdf_reader = PyPDF2.PdfReader(file)
            page = pdf_reader.pages[0]
        

            for line in page.extract_text().splitlines():
                if ("Total due by" in line):
                    invoiceTotal = line
                    invoiceTotal = invoiceTotal.split('$')
                    return invoiceTotal[1]
        

    except Exception as ex:
        print(f"Error while reading PDF: {ex}")
