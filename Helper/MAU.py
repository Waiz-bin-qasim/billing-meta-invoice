import pandas as pd
from Config.dbConfig import establish_connection, close_connection
import openpyxl
import datetime


def parseMAUFile(user):
    try:
        print(type(user))
        loc = "MAU.xlsx"
        wb = openpyxl.load_workbook(loc)
        ws = wb.active

        li = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,min_col=1, max_col=16, values_only=True):
            if all(cell is None for cell in row):
                break
            li.append(row)

        sqlQuery = "INSERT INTO mau (ORG_NAME_PRM,WABA_ID, MONTHLY_ACTIVE_USER_LIMIT, BAND_PRICE, AGENT_LICENCES, AGENT_LICENSE_PRICE, ADDITIONAL_AGENT_LICENSE_PRICE, SUPPORT_FEE,INV_MONTH, MAU_Count,INV_YEAR,ORG_PLAN,ADDITIONAL_CHARGES,DESCRIP,ADD_ONS,ADD_ONS_AGENT_PRICE) VALUES (%s,%s,%s,%s,%s,%s,%s, %s, %s, %s, %s, %s, %s,%s, %s,%s)"
        cursor, connection = establish_connection()
        if(li):
            month,year = getCredentials()
            cursor.executemany(sqlQuery, li)
            connection.commit()
            sqlquery = "INSERT INTO mau_logs(INV_MONTH,INV_YEAR, CREATED_BY, CREATED_ON) VALUES(%s,%s,%s,%s)"
            current_datetime = datetime.datetime.now()
            cursor.execute(sqlquery,(month,year,user,current_datetime))
            connection.commit()
            close_connection(cursor, connection)
            response = {
                'message': 'success'
            }
        else:
            response = {
                'message' : 'failed'
            }
            raise Exception("Failed to insert data in Excel File")
        return response
    except Exception as ex:
        close_connection(cursor, connection)
        print(f"error during inserting excel file: {ex}")
        return ex
    
    
def getCredentials():
    try:
        loc = "MAU.xlsx"
        wb = openpyxl.load_workbook(loc)
        ws = wb.active
        month = ws.cell(row=2, column=9).value
        year = ws.cell(row=2, column=11).value
        print(month)
        return month,year
    except Exception as ex:
        raise Exception(ex)
