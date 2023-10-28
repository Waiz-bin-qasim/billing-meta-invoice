from flask import Flask, render_template, request,jsonify,make_response,send_file,redirect,url_for,g, Blueprint
from flask_jwt_extended import JWTManager, jwt_required, get_jwt, get_jwt_identity, verify_jwt_in_request
from functools import wraps,partial
import jwt
import datetime
# from Middlewares.loginMiddlewares import token_required
import dataHandler
import getCsv
import os
from multiprocessing import Process
import threading
from Models.loginModels import loginCheck
from Helper.MAU import parseMAUFile
from Helper.loginHelpers import passwordDecrypt
from Helper.mauHelpers import checkMauLogs, getAllMau
from Helper.billingHelpers import getAllBilling, checkBillingLogs
from Helper.csvHelpers import generateCheck
from flask_socketio import SocketIO
from Sockets.sockets import showBar
from Sockets.sockets import updateProgress
from Helper.signup import addUser,displayUsers,displayRoles,delUser,updateUser
from Helper.dashboard import displayTotalClients,displayTotalInvoices,displayTotalUSD,displayTotalPKR,displayWhatsappAmount,displayBarChart
from flask_mail import Mail, Message
import secrets
from Helper.forgetPassword import confirmEmail,checkToken,setPassword
# month and year from MAU Billing
from Helper.MAU import getCredentials
# month and year from PDF Files
from oldPdfReader import getVariables as oldGetVariables
from newPdfReader import getVariables as newGetVariables
# for copying files
import shutil




# import computations

#initialising the server
IMG = os.path.join('static', 'Img')

app = Flask(__name__,template_folder='templates')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')  
app.config['ENCRYPT_KEY'] = os.environ.get('ENCRYPT_KEY')
# app.config['MAIL_SERVER'] = os.environ.get('MAIL_sERVER')
# app.config['MAIL_PORT'] = os.environ.get('MAIL_PORT')
# app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS')
# app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
# app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587  # Use the appropriate port for your email server
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'haziq.ahmed@eocean.com.pk'
app.config['MAIL_PASSWORD'] = 'mklv gqna kyou wuwx'
# socketio = SocketIO(app)
app.config['sideBarImage'] = IMG
mail = Mail(app)
def token_required(f):
    
    @wraps(f)
    def decorated(*args, **kwargs):
        
        try:
            if 'token' in request.cookies:
                token = request.cookies['token']
                data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"]) 
                
                # Extract the 'user' value from the decoded token
                print(data)
                user = data.get('user')
                permissions = data.get('permissions', [])
                role = data.get('role')
                print(role)
                if user is None:
                    return jsonify({'message': 'Invalid token: user not found'}), 401
                
                # Pass the 'user' as a keyword argument to the wrapped function
                kwargs['user'] = user
                kwargs['permissions'] = permissions
                kwargs['role'] = role
            else:
                return redirect(url_for("login"),code=307)
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': 'Invalid token'}), 401
        return f(*args, **kwargs)
    
    return decorated

def checkPermission(route,permissions):
    if(route in permissions):
        return True
    return False


# #token verified before executing
@app.route('/upload', methods=['POST',"GET"])
@token_required
def upload(user,permissions,role):

    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400 
        if request.method == "POST":
            updateProgress(socketio,"ascsac",20)
            data = request.form
            parserChoice = data.get('parserChoice')
            print(parserChoice)
            file = request.files['file']
            if file.filename == '':
                raise Exception('No file selected')
            fileName = 'transaction.pdf'
            file.save(fileName)
            sql_values = []
            print(user[0])
            updateProgress(socketio,"ascsac",35)
            if(checkBillingLogs(parserChoice) == True):
                updateProgress(socketio,"ascsac",50)
                response = dataHandler.run(sql_values,parserChoice,user,socketio)
                updateProgress(socketio,"ascsac",90)
                if response['message'] == 'failed':
                    return jsonify(response),400
    
                # place after parsing is confirmed
                if parserChoice == '0':
                    # old pdf reader
                    invoice_number, invoice_month, invoice_year = oldGetVariables(fileName)
                    file_path = f'./metaInvoiceFiles/{invoice_month}{invoice_year}.pdf'
                    shutil.copy(fileName, file_path)

                elif parserChoice == '1':
                    # new pdf reader
                    invoice_number, invoice_month, invoice_year = newGetVariables(fileName)
                    file_path = f'./metaInvoiceFiles/{invoice_month}{invoice_year}'
                    shutil.copy(fileName, file_path)

                return jsonify(response),200
            else:
                updateProgress(socketio,"ascsac",90)
                return jsonify({'message': 'Meta Invoice Already Exists'}), 400
        else:
            data = getAllBilling()
            return jsonify(data)
    except Exception as ex:
        print(f'Error during file upload: {ex}')
        return jsonify({'message': 'An error occurred during file upload.'}), 400


#for making token
@app.route('/',methods = ['POST'])
def login():
    
    try: 
        
        if request.method == "POST":
            authUsername = [request.form['username']]
            authPassword = [request.form["password"]]
            print(authPassword)
            token,roleName = loginCheck(authUsername,authPassword)
            if token != 0:
                return jsonify({'token' : token,"roleName":roleName}),200   
            else:
               return jsonify({"message":"Incorrect Credentials or Account not Active"}),401
            
    except (NameError, TypeError) as error:
     print(error)
     return jsonify({error}), 400

@app.route('/downloadcsv', methods = ['GET'])
@token_required
def downloadcsv(user,permissions,role):
    try:
            route = request.endpoint
            if(checkPermission(route,permissions) == False):
                return jsonify({'message': 'Permission Not Given'}), 400
            fileName = os.listdir("excel/")
            response = []
            count = 0
            for name in fileName:
                updatedtAt = datetime.datetime.fromtimestamp(os.path.getmtime("./excel/"+name))
                updatedtAt = updatedtAt.strftime('%B %d, %Y ')
                print(updatedtAt)
                createdAt = datetime.datetime.fromtimestamp(os.path.getctime("./excel/"+name))
                createdAt = createdAt.strftime('%B %d, %Y ')
                response.append([count+1,name.split('.')[0],createdAt,updatedtAt]) 
                count +=1
            return jsonify(response)
    except Exception as ex:
        print(f"Error during file download: {ex}")
        return jsonify({'Error Ocurred' : ex}), 500  

@app.route('/mau/upload',methods = ['POST',"GET"])
@token_required
def mau(user,permissions,role):
    try:
         route = request.endpoint
         if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
         if request.method == "POST":
            updateProgress(socketio,"a",20)
            print(user[0])
            data = request.form
            file = request.files['file']
            if file.filename == '':
                raise Exception('No file selected')
            
            fileName = 'MAU.xlsx'
            file.save(fileName)

            # place after parsing is confirmed
            month, year = getCredentials()
            file_path = f'./billingMAUFiles/{month}{year}.xlsx'
            shutil.copy(fileName, file_path)

            updateProgress(socketio,"a",30)
            if (checkMauLogs() == True):
                updateProgress(socketio,"a",40)
                response = parseMAUFile(user,socketio)
                updateProgress(socketio,"a",80)
                return jsonify(response)
            else:
                updateProgress(socketio,"a",80)
                return jsonify({'error': 'Bad Request'}), 400
            
         else:
             data = getAllMau()
             return jsonify(data)
    except Exception as ex:
     print(ex)
     print(f'Error during upload: {ex}')
     return jsonify({'Error Ocurred' : ex}), 400


@app.route('/generatecsv/<socketId>',methods = ['POST'])
@token_required
def generateCsv(user,permissions,socketId,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        param1 = request.args.get('param1')
        param2 = request.args.get('param2')
        if param1 and param2:
            updateProgress(socketio,socketId,20)
            response = generateCheck(param1,param2)
            if(response == True):
                updateProgress(socketio,socketId,40)
                file_path = getCsv.run(param1, param2,socketio,socketId)
                g.file_path = file_path
                updateProgress(socketio,socketId,100)
                return jsonify("File was Generated")
            else:
                updateProgress(socketio,socketId,40)
                return jsonify(response),400
        else:
            return jsonify({"message":"Please provide both param1 and param2 as query parameters."}),400
    except Exception as e:
        # print(f'Error during generating file: {e}')
        return jsonify({'Error Ocurred': e}), 400 

@app.route('/getcsv', methods = ['GET'])
@token_required
def getcsv(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        param1 = request.args.get('param1')
        param2 = request.args.get('param2')
        if param1 and param2:
            file_path = "excel/"+param1+param2+".xlsx"
            return send_file(file_path, as_attachment=True, download_name=f'{param1+param2}.xlsx'),200
        else:
            raise Exception("Please provide both param1 and param2 as query parameters.")
    except Exception as e:
        print(f"Error during file download: {e}")
        return jsonify({'Error Ocurred' : e}), 400 





@app.route('/getpdf', methods = ['GET'])
@token_required
def getpdf(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        
        param1 = request.args.get('param1')
        param2 = request.args.get('param2')
        if param1 and param2: 
            file_path = "metaInvoiceFiles/"+param1+param2+".pdf"
            return send_file(file_path, as_attachment=True, download_name=f'{param1 + param2}.pdf'),200
        else: 
            raise Exception("Please provide both param1 and param2")
        
    except Exception as e:
        print("Error occured: ", {e})
        return jsonify({"Error Occured " : e}), 400        



@app.route('/getmau', methods = ['GET'])
@token_required
def getmau(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
    
        param1 = request.args.get('param1')
        param2 = request.args.get('param2')
        if param1 and param2: 
            file_path = "billingMAUFiles/"+param1+param2+".xlsx"
            return send_file(file_path, as_attachment=True, download_name=f'{param1 + param2}.xlsx'),200
        else: 
            raise Exception("Please provide both param1 and param2")
        
    except Exception as e:
        print("Error occured: ", {e})
        return jsonify({"Error Occured " : e}), 400   
    


@app.route('/finance/upload',methods = ['POST','GET'])
@token_required
def financeUpload(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        
        if request.method == "POST":
            file = request.files['file']
            if file.filename == '':
                raise Exception('No file selected')
            

            fileName = 'financeReport.xlsx'
            file.save(fileName)

            # <getting month and year>
            import openpyxl
            wb = openpyxl.load_workbook(fileName)
            ws = wb.active
            month = ws.cell(row=3, column=1).value
            year = ws.cell(row=3, column=1).value
            month = month[:3]
            year = year[4:]


            # storing it in folder
            file_path = f'./financeReportFiles/{month}{year}.xlsx'
            shutil.copy(fileName, file_path)
            return jsonify({'message': 'File uploaded successfully'}),200
        else:
            param1 = request.args.get('param1')
            param2 = request.args.get('param2')
            if param1 and param2: 
                file_path = "financeReportFiles/"+param1+param2+".xlsx"
                return send_file(file_path, as_attachment=True, download_name=f'{param1 + param2}.xlsx'),200
            else: 
                raise Exception("Please provide both param1 and param2")
    
    except Exception as ex:
     print(ex)
     print(f'Error during upload: {ex}')
     return jsonify({'Error Occured' : ex}), 400
    



@app.route('/displayrole',methods = ['GET'])
@token_required
def displayRole(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        data = displayRoles()
        return jsonify(data)
    except Exception as ex:
        return jsonify({'message':'error during displaying roles'}),400

@app.route('/displaydashboard',methods = ['GET'])
@token_required
def displayDashboard(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        month = request.args.get('month')
        year = request.args.get('year')
        data1 = displayTotalUSD(month,year)
        
        data2 = displayTotalPKR(month,year)
        
        data3 = displayWhatsappAmount(month,year)
        
        data4 = displayWhatsappAmount(month,year)
        
        data5 = displayTotalClients(month,year)
        
        data6 = displayTotalInvoices()
        
        data7 = displayBarChart()
        return jsonify(data1,data2,data3,data4,data5,data6,data7)

        
    except Exception as ex:
        return jsonify({'message':'error during displaying dashboard'}),400

# @app.route('/displaypkr',methods = ['GET'])
# @token_required
# def displayPKR(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         month = request.args.get('month')
#         year = request.args.get('year')
#         data = displayTotalPKR(month,year)
#         return jsonify(data)
#     except Exception as ex:
#         return jsonify({'message':'error during displaying pkr amount'}),400

# @app.route('/displaywhatsapp',methods = ['GET'])
# @token_required
# def displayWhatsapp(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         month = request.args.get('month')
#         year = request.args.get('year')
#         data = displayWhatsappAmount(month,year)
#         return jsonify(data)
#     except Exception as ex:
#         return jsonify({'message':'error during displaying usd amount'}),400

# @app.route('/displaytotal',methods = ['GET'])
# @token_required
# def displayTotal(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         month = request.args.get('month')
#         year = request.args.get('year')
#         usdEarned = displayTotalUSD(month,year)
#         usdWhatsapp = displayWhatsappAmount(month,year)
#         # usdEarned = usdEarned.replace(',', '').strip()
#         usdWhatsapp = float(usdWhatsapp.replace(',', '').strip())
#         data = usdEarned + usdWhatsapp
#         return jsonify(data)
#     except Exception as ex:
#         print(ex)
#         return jsonify({'error':'error during total'}),400



@app.route('/forgetpassword',methods = ['POST'])
def forgetPassword():
    try:
        email = request.form['email']
        email = [email]
        token = confirmEmail(email)
        
        print(type(email))
        print(token)
        msg = Message('Password Reset', sender='haziq.ahmed@eocean.com.pk', recipients=email)
        msg.body = f'Code to reset your password: {token}'
        mail.send(msg)
        return jsonify({'message':'success'})
    except Exception as ex:
        print(ex)
        return jsonify({'message':'error during forget password'}),400


@app.route('/resetpassword',methods = ['POST'])
def resetPassword():
    try:
    
        email = request.form['email']
        newPassword = request.form['newPassword']
        confirmPassword = request.form['confirmPassword']
        if(newPassword == confirmPassword):
            response = setPassword(email,newPassword)
            if(response == 0):
                return jsonify({'message':'failed'}),400
        else:
            return jsonify({'message': 'passwords does not match'}),400
        
        return jsonify({'message':'success'}),200
    except Exception as ex:
        print(ex)
        return jsonify({'message':'error during forget password'}),400
    
@app.route('/changepassword',methods = ['POST'])
@token_required
def changePassword(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        email = request.form['email']
        newPassword = request.form['newPassword']
        confirmPassword = request.form['confirmPassword']
        if(newPassword == confirmPassword):
            response = setPassword(email,newPassword)
            if(response == 1):
                return jsonify({'message' : 'success'})
            return jsonify({'message':'failed'}),400
        else:
            return jsonify({'message': 'passwords does not match'}),400
        
    except Exception as ex:
        print(ex)
        return jsonify({'message':'error during forget password'}),400

@app.route('/entertoken',methods = ['POST'])
def enterToken():
    try:
        email = request.form['email']
        token = request.form['token']
        result = checkToken(email,token)
        if(result):
            return jsonify({'message':'success'})
        else:
            return jsonify({'message':'failed'}),400
    except Exception as ex:
        print(ex)
        return jsonify({'message':'token incorrect'}),400

# @app.route('/displayclients',methods = ['GET'])
# @token_required
# def displayClients(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         month = request.args.get('month')
#         year = request.args.get('year')
#         data = displayTotalClients(month,year)
#         return jsonify(data)
#     except Exception as ex:
#         print(ex)
#         return jsonify({'message':'error during displaying clients'}),400

# @app.route('/displayinvoice',methods = ['GET'])
# @token_required
# def displayInvoice(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         data = displayTotalInvoices()
#         return jsonify(data)
#     except Exception as ex:
#         return jsonify({'message':'error during displaying invoices'}),400
    
# @app.route('/displaychart',methods = ['GET'])
# @token_required
# def displayChart(user,permissions,role):
#     try:
#         route = request.endpoint
#         if(checkPermission(route,permissions) == False):
#             return jsonify({'message': 'Permission Not Given'}), 400
#         data = displayBarChart()
#         return jsonify(data)
#     except Exception as ex:
#         return jsonify({'message':'error during displaying Bar Chart Data'}),400

@app.route('/adduser',methods = ['POST','GET','DELETE','PATCH'])
@token_required
def addUserData(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        if request.method == "POST":
            userData = request.form
            userFirstName = userData["firstName"]
            userLastName = userData["lastName"]
            userEmail = userData["email"]
            userPassword = userData["password"]
            userRoleId = userData["roleId"]
            response = addUser(userFirstName,userLastName,userEmail,userPassword,userRoleId)
            return jsonify(response)
        elif request.method == "DELETE":
            user = request.args.get('user')
            response = delUser(user)
            return response
        elif request.method == 'PATCH':
            username = request.form['username']
            columns = request.form.getlist('columns[]')  
            values = request.form.getlist('values[]') 
            # columns = request.args.get['username']
            # values = request.args.get['values']
            if len(columns) != len(values):
                return jsonify({'error': 'Number of columns and values must match'}),400
            updateUser(username,columns,values)
            return jsonify({'message': 'Update successful'})
        else:
            data = displayUsers()
            return jsonify(data)
            
    except Exception as ex:
        print(ex)
        return jsonify({'error in user module': str(ex)}),400    
    
    

@app.route('/finance/reports',methods = ['get'])
@token_required
def FinanceReport(user,permissions,role):
    try:
        route = request.endpoint
        if(checkPermission(route,permissions) == False):
            return jsonify({'message': 'Permission Not Given'}), 400
        
        fileName = os.listdir("financeReportFiles/")
        response = []
        count = 0
        for name in fileName:
            updatedtAt = datetime.datetime.fromtimestamp(os.path.getmtime("./excel/"+name))
            updatedtAt = updatedtAt.strftime('%B %d, %Y ')
            print(updatedtAt)
            createdAt = datetime.datetime.fromtimestamp(os.path.getctime("./excel/"+name))
            createdAt = createdAt.strftime('%B %d, %Y ')
            response.append([count+1,name.split('.')[0],updatedtAt,createdAt]) 
            count +=1
        flag = False
        if(role == 'admin'):
            flag = True
        return jsonify(response)
    except Exception as ex:
        print(f"Error during file download: {ex}")
        return jsonify({'Error Ocurred' : ex}), 500  



#server starting
if __name__ == '__main__':
    app.run(host='0.0.0.0',port=8090,debug=True)
